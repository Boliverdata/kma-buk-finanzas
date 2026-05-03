#!/usr/bin/env python3
"""
Extractor diario de datos financieros desde Buk Finanzas (ex Bemmbo) para KMA Asset Management.

- Egresos (facturas recibidas) para todas las empresas, con categoría
- Ingresos (facturas emitidas) solo para I-DEAL
- Valorización en UF para fechas de emisión y agendamiento
- Sube el consolidado a Google Drive
"""

import os
import json
import sys
import time
import logging
from datetime import datetime, date
from pathlib import Path

import requests
import pandas as pd
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google.oauth2.service_account import Credentials

# ──────────────────────────────────────────────────────────────
# Configuración
# ──────────────────────────────────────────────────────────────
BUK_BASE          = "https://api.bemmbo.com/v1"
HISTORICAL_FROM   = "2025-01-01"
DRIVE_FOLDER_ID   = "1h8lS7xFVMwciI2fOHQXIuhMMNF075dDo"
PAGE_SIZE         = 500
IDLEAL_KEY        = "I-DEAL"

TOKENS_XLSX       = Path(__file__).parent / "Token Buk.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────
# Carga de empresas y tokens
# ──────────────────────────────────────────────────────────────
def load_companies() -> dict[str, str]:
    """
    1. Intenta leer desde la variable de entorno COMPANY_TOKENS (JSON).
    2. Si no existe, carga desde Token Buk.xlsx (columnas: Nombre de la Empresa, Token).
    """
    raw = os.environ.get("COMPANY_TOKENS")
    if raw:
        return json.loads(raw)

    if TOKENS_XLSX.exists():
        df = pd.read_excel(TOKENS_XLSX)
        col_empresa = next(c for c in df.columns if "empresa" in c.lower() or "nombre" in c.lower())
        col_token   = next(c for c in df.columns if "token" in c.lower())
        companies = {
            str(row[col_empresa]).strip(): str(row[col_token]).strip()
            for _, row in df.iterrows()
            if str(row[col_token]).strip().startswith("eyJ")
        }
        log.info(f"Tokens cargados desde {TOKENS_XLSX.name}: {len(companies)} empresas")
        return companies

    raise EnvironmentError(
        "No se encontraron tokens. Define COMPANY_TOKENS o coloca Token Buk.xlsx junto al script."
    )


# ──────────────────────────────────────────────────────────────
# API UF — mindicador.cl (CMF)
# Carga el año completo de una vez (1 request por año en vez de 1 por fecha)
# ──────────────────────────────────────────────────────────────
_uf_cache: dict[str, float] = {}
_uf_years_loaded: set[int] = set()


def _load_uf_year(year: int) -> None:
    """Precarga todos los valores UF de un año en el caché."""
    if year in _uf_years_loaded:
        return
    try:
        resp = requests.get(f"https://mindicador.cl/api/uf/{year}", timeout=20)
        resp.raise_for_status()
        for item in resp.json().get("serie", []):
            day = str(item["fecha"])[:10]
            _uf_cache[day] = float(item["valor"])
        _uf_years_loaded.add(year)
        log.info(f"UF {year}: {len([k for k in _uf_cache if k.startswith(str(year))])} valores cargados")
    except Exception as exc:
        log.warning(f"No se pudo precargar UF año {year}: {exc}")


def get_uf(date_str: str | None) -> float | None:
    """Devuelve el valor de la UF para una fecha YYYY-MM-DD."""
    if not date_str:
        return None
    date_str = str(date_str)[:10]
    year = int(date_str[:4])
    _load_uf_year(year)
    return _uf_cache.get(date_str)


# ──────────────────────────────────────────────────────────────
# API Buk Finanzas — helpers
# ──────────────────────────────────────────────────────────────
def buk_headers(token: str) -> dict:
    return {"Authorization": token, "Accept": "application/json"}


def fetch_all_pages(url: str, token: str, extra_params: dict | None = None) -> list[dict]:
    """Itera todas las páginas de un endpoint (500 registros/página)."""
    all_records: list[dict] = []
    page = 0
    base_params = {
        "options_pageSize": PAGE_SIZE,
        "options_sortBy":   "DATE",
        "options_sortOrder": "ASC",
        **(extra_params or {}),
    }
    while True:
        params = {**base_params, "options_page": page}
        try:
            resp = requests.get(url, headers=buk_headers(token), params=params, timeout=30)
            resp.raise_for_status()
            body = resp.json()
        except Exception as exc:
            log.error(f"  Error en {url} página {page}: {exc}")
            break

        records: list = (
            body if isinstance(body, list)
            else body.get("pageResults") or body.get("data") or body.get("items") or body.get("results") or []
        )
        if not records:
            break
        all_records.extend(records)
        log.info(f"    Página {page}: {len(records):>4} registros  (acum: {len(all_records):>5})")
        if len(records) < PAGE_SIZE:
            break
        page += 1
        time.sleep(0.25)

    return all_records


# ──────────────────────────────────────────────────────────────
# Extracción de campos
# ──────────────────────────────────────────────────────────────
def extract_categoria(inv: dict) -> str:
    """
    Extrae la(s) categoría(s) del egreso.
    Usa el array 'categories' (campo simple) o 'accountingEntries[].category.name'.
    Si hay múltiples categorías, las une con ' | '.
    """
    nombres = []

    # Fuente 1: array categories[] (más directo)
    for cat in inv.get("categories") or []:
        if isinstance(cat, dict):
            n = cat.get("name") or cat.get("accountingName")
        else:
            n = str(cat)
        if n:
            nombres.append(str(n))

    # Fuente 2: accountingEntries[].category.name (fallback)
    if not nombres:
        for entry in inv.get("accountingEntries") or []:
            if isinstance(entry, dict):
                cat = entry.get("category") or {}
                n = cat.get("name") or cat.get("accountingName") if isinstance(cat, dict) else None
                if n:
                    nombres.append(str(n))

    return " | ".join(dict.fromkeys(nombres))  # únicos, orden preservado


def extract_payment_date(inv: dict) -> str | None:
    """Fecha en que se pagó o se programó el pago."""
    for field in ["paidBy", "paidAt", "paymentDate", "paidDate"]:
        v = inv.get(field)
        if v:
            return str(v)[:10]
    for p in inv.get("payments") or []:
        v = p.get("scheduledPaymentDate")
        if v:
            return str(v)[:10]
    for field in ["paymentScheduledAt", "scheduledPaymentDate", "scheduledAt"]:
        v = inv.get(field)
        if v:
            return str(v)[:10]
    return None


def extract_first_date(inv: dict) -> str | None:
    """Fecha de emisión/recepción del documento."""
    for field in ["emittedAt", "issuedAt", "receivedDate", "date", "emissionDate", "createdAt"]:
        v = inv.get(field)
        if v:
            return str(v)[:10]
    return None


def normalize_invoice(inv: dict, empresa: str, tipo_flujo: str) -> dict:
    """Convierte un dict crudo de Buk Finanzas en una fila normalizada."""
    counterpart = inv.get("issuer") or inv.get("customer") or inv.get("supplier") or {}
    if not isinstance(counterpart, dict):
        counterpart = {}

    rut = (
        inv.get("issuerFiscalId")
        or inv.get("customerFiscalId")
        or inv.get("supplierFiscalId")
        or counterpart.get("fiscalId")
        or counterpart.get("rut")
        or ""
    )
    nombre = (
        inv.get("issuerName")
        or inv.get("customerName")
        or inv.get("supplierName")
        or counterpart.get("name")
        or counterpart.get("legalName")
        or counterpart.get("businessName")
        or ""
    )

    fecha_emision      = extract_first_date(inv)
    fecha_agendamiento = extract_payment_date(inv)

    monto_neto  = inv.get("netAmount")   or inv.get("net")   or inv.get("subtotal") or 0
    monto_bruto = inv.get("totalAmount") or inv.get("total") or inv.get("amount")   or 0

    uf_emision      = get_uf(fecha_emision)
    uf_agendamiento = get_uf(fecha_agendamiento)

    def safe_div(monto, uf):
        if uf and uf > 0 and monto:
            return round(float(monto) / float(uf), 4)
        return None

    categoria = extract_categoria(inv) if tipo_flujo == "EGRESO" else ""

    return {
        "empresa":               empresa,
        "tipo_flujo":            tipo_flujo,
        "categoria":             categoria,
        "rut_contraparte":       rut,
        "nombre_contraparte":    nombre,
        "fecha_emision":         fecha_emision,
        "numero_documento":      inv.get("number") or inv.get("invoiceNumber") or inv.get("folio") or inv.get("id") or "",
        "tipo_documento":        inv.get("type")   or inv.get("documentType") or "",
        "fecha_agendamiento":    fecha_agendamiento,
        "monto_neto":            monto_neto,
        "monto_bruto":           monto_bruto,
        "estado":                inv.get("status") or inv.get("state") or "",
        "uf_dia_emision":        uf_emision,
        "uf_dia_agendamiento":   uf_agendamiento,
        "monto_uf_emision":      safe_div(monto_bruto, uf_emision),
        "monto_uf_agendamiento": safe_div(monto_bruto, uf_agendamiento),
    }


# ──────────────────────────────────────────────────────────────
# Modo descubrimiento (--discover)
# ──────────────────────────────────────────────────────────────
def discover_fields(companies: dict):
    empresa, token = next(iter(companies.items()))
    log.info(f"[DISCOVER] Empresa: {empresa}")
    resp = requests.get(
        f"{BUK_BASE}/invoices/received",
        headers=buk_headers(token),
        params={"options_page": 0, "options_pageSize": 1},
        timeout=30,
    )
    resp.raise_for_status()
    print("\n──── RAW RESPONSE (primer egreso) ────")
    print(json.dumps(resp.json(), indent=2, ensure_ascii=False, default=str))
    print("─────────────────────────────────────\n")
    sys.exit(0)


# ──────────────────────────────────────────────────────────────
# Extracción por empresa
# ──────────────────────────────────────────────────────────────
def extract_company(empresa: str, token: str, include_ingresos: bool = False) -> list[dict]:
    rows: list[dict] = []

    log.info(f"  ► Egresos — {empresa}")
    for inv in fetch_all_pages(f"{BUK_BASE}/invoices/received", token, {"receivedDateSince": HISTORICAL_FROM}):
        rows.append(normalize_invoice(inv, empresa, "EGRESO"))

    if include_ingresos:
        log.info(f"  ► Ingresos — {empresa}")
        for inv in fetch_all_pages(f"{BUK_BASE}/invoices/issued", token, {"emissiondateSince": HISTORICAL_FROM}):
            rows.append(normalize_invoice(inv, empresa, "INGRESO"))

    log.info(f"  ✓ {empresa}: {len(rows):,} filas")
    return rows


# ──────────────────────────────────────────────────────────────
# Google Drive
# ──────────────────────────────────────────────────────────────
def get_drive_service():
    creds_info = json.loads(os.environ["GOOGLE_CREDENTIALS"])
    creds = Credentials.from_service_account_info(
        creds_info,
        scopes=["https://www.googleapis.com/auth/drive"],
    )
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def upload_to_drive(local_path: str, file_name: str):
    service = get_drive_service()
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    existing = (
        service.files()
        .list(
            q=f"name='{file_name}' and '{DRIVE_FOLDER_ID}' in parents and trashed=false",
            fields="files(id,name)",
        )
        .execute()
        .get("files", [])
    )

    media = MediaFileUpload(local_path, mimetype=mime, resumable=True)
    if existing:
        service.files().update(fileId=existing[0]["id"], media_body=media).execute()
        log.info(f"Drive: archivo actualizado → {file_name}")
    else:
        meta = {"name": file_name, "parents": [DRIVE_FOLDER_ID]}
        service.files().create(body=meta, media_body=media, fields="id").execute()
        log.info(f"Drive: archivo creado → {file_name}")


# ──────────────────────────────────────────────────────────────
# Generación de Excel
# ──────────────────────────────────────────────────────────────
COLUMN_WIDTHS = {
    "empresa": 18, "tipo_flujo": 10, "categoria": 28,
    "rut_contraparte": 14, "nombre_contraparte": 35,
    "fecha_emision": 14, "numero_documento": 16, "tipo_documento": 22,
    "fecha_agendamiento": 18, "monto_neto": 14, "monto_bruto": 14,
    "estado": 14, "uf_dia_emision": 14, "uf_dia_agendamiento": 18,
    "monto_uf_emision": 16, "monto_uf_agendamiento": 20,
}

NUMBER_COLS = {
    "monto_neto", "monto_bruto",
    "uf_dia_emision", "uf_dia_agendamiento",
    "monto_uf_emision", "monto_uf_agendamiento",
}


def save_excel(df: pd.DataFrame, local_path: str):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(local_path, engine="openpyxl", date_format="DD-MM-YYYY") as writer:
        df.to_excel(writer, index=False, sheet_name="Consolidado")
        ws = writer.sheets["Consolidado"]

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for i, col_name in enumerate(df.columns, start=1):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = COLUMN_WIDTHS.get(col_name, 16)
            if col_name in NUMBER_COLS:
                for cell in ws.iter_cols(min_col=i, max_col=i, min_row=2):
                    for c in cell:
                        c.number_format = "#,##0.00"

    log.info(f"Excel guardado: {local_path}  ({len(df):,} filas)")


# ──────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────
def main():
    companies = load_companies()

    if "--discover" in sys.argv:
        discover_fields(companies)

    all_rows: list[dict] = []
    for empresa, token in companies.items():
        log.info(f"━━━ {empresa} ━━━")
        rows = extract_company(
            empresa,
            token,
            include_ingresos=(empresa.upper() == IDLEAL_KEY.upper()),
        )
        all_rows.extend(rows)

    log.info(f"Total filas extraídas: {len(all_rows):,}")

    df = pd.DataFrame(all_rows)
    for col in ["fecha_emision", "fecha_agendamiento"]:
        df[col] = pd.to_datetime(df[col], errors="coerce").dt.date

    today      = date.today().strftime("%Y%m%d")
    file_name  = f"{today}_Consolidado_Buk_Finanzas.xlsx"
    local_path = f"/tmp/{file_name}"

    save_excel(df, local_path)
    upload_to_drive(local_path, file_name)
    log.info("✓ Proceso completado.")


if __name__ == "__main__":
    main()
